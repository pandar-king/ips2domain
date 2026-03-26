import re
import sys
import socket
import random
import getopt
import requests
import openpyxl
import time
import json
import os
import signal
from urllib.parse import urlparse
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor

# 全局变量用于保存状态
STATE_FILE = ".revip_state.json"
save_state_flag = False
current_state = {}


HIGHLIGHT_FILL = PatternFill(start_color='FFFF00', fill_type='solid')
HEADER_FILL = PatternFill(start_color='DDDDDD', fill_type='solid')
VERSION = "version1.0"
AUTHOR = "panda"

# 全局代理配置
PROXY = {'http': None, 'https': None}

# 批量查询计数器
BATCH_SIZE = 50

def save_state():
    """保存当前进度到状态文件"""
    global current_state
    if current_state:
        try:
            with open(STATE_FILE, 'w', encoding='utf-8') as f:
                json.dump(current_state, f, ensure_ascii=False, indent=2)
            print(f"\n💾 进度已保存到 {STATE_FILE}")
        except Exception as e:
            print(f"\n⚠️ 保存进度失败: {e}")

def load_state():
    """从状态文件加载进度"""
    global current_state
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE, 'r', encoding='utf-8') as f:
                current_state = json.load(f)
            return current_state
        except Exception as e:
            print(f"⚠️ 加载进度失败: {e}")
    return None

def clear_state():
    """清除状态文件"""
    global current_state
    if os.path.exists(STATE_FILE):
        try:
            os.remove(STATE_FILE)
            current_state = {}
        except Exception:
            pass

def signal_handler(sig, frame):
    """处理中断信号"""
    global save_state_flag
    print("\n\n⚠️ 检测到中断，正在保存进度...")
    save_state_flag = True

# 注册信号处理函数
signal.signal(signal.SIGINT, signal_handler)
signal.signal(signal.SIGTERM, signal_handler)

def print_banner():
    banner = f"""
╔══════════════════════════════════════════════════════════════════╗
║                                                                  ║
║   ██╗██████╗ ███████╗    ██████╗  ██████╗ ███╗   ███╗ █████╗ ██╗███╗   ██╗  ║
║   ██║██╔══██╗██╔════╝    ██╔══██╗██╔═══██╗████╗ ████║██╔══██╗██║████╗  ██║  ║
║   ██║██████╔╝███████╗    ██║  ██║██║   ██║██╔████╔██║███████║██║██╔██╗ ██║  ║
║   ██║██╔═══╝ ╚════██║    ██║  ██║██║   ██║██║╚██╔╝██║██╔══██║██║██║╚██╗██║  ║
║   ██║██║     ███████║    ██████╔╝╚██████╔╝██║ ╚═╝ ██║██║  ██║██║██║ ╚████║  ║
║   ╚═╝╚═╝     ╚══════╝    ╚═════╝  ╚═════╝ ╚═╝     ╚═╝╚═╝  ╚═╝╚═╝╚═╝  ╚═══╝  ║
║                                                                  ║
║                    IP反查域名工具 {VERSION}                        ║
║                         by {AUTHOR}                              ║
║                                                                  ║
╚══════════════════════════════════════════════════════════════════╝
"""
    print(banner)

def clean_target(target):
    """智能清洗输入目标"""
    try:
        target = target.strip(" '\"")
        if not target.startswith(('http://', 'https://')):
            target = f'http://{target}'
        
        parsed = urlparse(target)
        hostname = parsed.hostname
        if not hostname:
            return None
        
        return socket.gethostbyname(hostname)
    except Exception as e:
        print(f"解析失败: {str(e)}")
        return None

def user_agents():
    """国内主流浏览器User-Agent"""
    return [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:83.0) Gecko/20100101 Firefox/83.0",
        "Mozilla/5.0 (Linux; Android 10; M2007J3SC) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Mobile Safari/537.36",
        "Mozilla/5.0 (iPhone; CPU iPhone OS 14_2 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0.1 Mobile/15E148 Safari/604.1"
    ]

def fetch_domains_cn(ip, use_proxy=True, max_retries=3):
    """使用国内接口查询IP关联域名，支持失败重试"""
    global PROXY
    headers = {
        'User-Agent': random.choice(user_agents()),
        'Referer': 'https://site.ip138.com/'
    }
    """国内接口可配"""
    apis = [
        {
            'url': f'https://site.ip138.com/{ip}/',
            'method': 'regex',
            'pattern': r'<li><span class="date">.*?</span><a href="/(.*?)/" target="_blank">'
        },
        #站点关闭
        # {
        #     'url': f'https://api.webscan.cc/?query={ip}',
        #     'method': 'json',
        #     'field': 'domain'
        # }
    ]
    
    domains = []
    for api in apis:
        retry_count = 0
        while retry_count < max_retries:
            try:
                session = requests.Session()
                session.trust_env = False
                # 根据use_proxy参数决定是否使用代理
                proxies = PROXY if use_proxy else {'http': None, 'https': None}
                response = session.get(
                    api['url'],
                    headers=headers,
                    timeout=15,
                    proxies=proxies
                )
                
                if response.status_code != 200:
                    retry_count += 1
                    if retry_count < max_retries:
                        wait_time = random.uniform(3, 6)
                        time.sleep(wait_time)
                    continue
                    
                if api['method'] == 'regex':
                    matches = re.findall(api['pattern'], response.text)
                    cleaned = [m.strip() for m in matches if m.strip()]
                    domains.extend(cleaned)
                    
                elif api['method'] == 'json':
                    data = response.json()
                    if isinstance(data, list):
                        valid = [str(d.get(api['field'], '')).strip() for d in data]
                        domains.extend([v for v in valid if v])
                
                # 去重并限制最大数量防止站点重定向
                domains = list(set(domains))[:50]
                time.sleep(random.uniform(1, 2))
                break  # 成功则跳出重试循环
                
            except Exception as e:
                retry_count += 1
                if retry_count < max_retries:
                    wait_time = random.uniform(3, 6)
                    time.sleep(wait_time)
                else:
                    print(f"接口 {api['url']} 查询失败（已重试{max_retries}次）: {str(e)}")
            
    return domains

# 全局查询计数器
query_counter = {'current': 0, 'total': 0}

def process_target(target, use_proxy=True):
    """处理目标"""
    global query_counter
    
    # 增加计数
    query_counter['current'] += 1
    current = query_counter['current']
    total = query_counter['total']
    
    print(f"\n[ {current}/{total} ] 正在查询: {target}")
    
    ip = clean_target(target)
    if not ip:
        print(f"❌ 目标解析失败: {target}")
        print("-" * 50)
        return (target, None)
    
    domains = fetch_domains_cn(ip, use_proxy)
    
    original_host = urlparse(target).hostname or target.split('//')[-1].split('/')[0]
    highlighted_domains = [
        f"\033[93m{d}*\033[0m" if (ip in d or original_host in d) else d 
        for d in domains if isinstance(d, str)
    ]
    
    ip_display = f"\033[92m{ip}\033[0m" if domains else ip
    
    print(f"► 原始输入: \033[94m{target}\033[0m")
    print(f"► 解析IP : {ip_display}")
    print(f"► 关联域名: {len(domains)} 个")
    
    if domains:
        print("  " + "\n  ".join(highlighted_domains))
    else:
        print("  未找到关联域名")
    
    print("-" * 50) 
    
    return (target, {"ip": ip, "domains": domains})

def extract_main_domain(domain):
    """从域名中提取主域名"""
    if not domain or not isinstance(domain, str):
        return ""
    
    # 移除可能的端口号
    domain = domain.split(':')[0]
    
    # 按点分割
    parts = domain.split('.')
    
    # 如果域名部分少于2个，直接返回原域名
    if len(parts) < 2:
        return domain
    
    # 常见顶级域名后缀列表
    tld_list = [
        'com', 'cn', 'net', 'org', 'gov', 'edu', 'mil', 'int',
        'co', 'io', 'cc', 'tv', 'biz', 'info', 'name', 'pro',
        'aero', 'asia', 'cat', 'jobs', 'mobi', 'museum', 'travel',
        'arpa', 'root', 'localhost'
    ]
    
    # 处理双后缀的情况 (如 .com.cn, .co.uk 等)
    if len(parts) >= 3:
        # 检查后两部分是否是组合后缀
        second_tld = parts[-2].lower()
        if second_tld in ['com', 'co', 'org', 'net', 'gov', 'edu', 'mil']:
            return '.'.join(parts[-3:])
    
    # 默认返回最后两部分
    return '.'.join(parts[-2:])

def query_icp_info(domain, use_proxy=True):
    """查询域名ICP备案信息"""
    global PROXY
    try:
        session = requests.Session()
        session.trust_env = False
        
        # 获取当前代理配置
        proxies = PROXY if use_proxy else {'http': None, 'https': None}
        
        url = f"https://uapis.cn/api/v1/network/icp?domain={domain}"
        response = session.get(url, timeout=10, proxies=proxies)
        
        if response.status_code == 200:
            data = response.json()
            if data.get('code') == '200' or data.get('code') == 200:
                service_licence = data.get('serviceLicence', '')
                unit_name = data.get('unitName', '')
                if service_licence and unit_name:
                    return f"{service_licence}-{unit_name}"
        return ""
    except Exception as e:
        return ""

def query_icp_with_cache(domain, use_proxy=True, icp_cache=None):
    """查询ICP备案信息，支持缓存"""
    global save_state_flag
    
    # 检查是否收到中断信号
    if save_state_flag:
        return None
    
    # 检查缓存
    if icp_cache and domain in icp_cache:
        return icp_cache[domain]
    
    icp_info = query_icp_info(domain, use_proxy)
    
    # 保存到缓存
    if icp_cache is not None:
        icp_cache[domain] = icp_info
    
    return icp_info

def export_results_with_progress(results, filename, use_proxy=True, icp_cache=None):
    """导出Excel，支持进度显示和中断恢复，单条IP主域名去重"""
    global save_state_flag, current_state
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "反查结果"
    
    headers = ["原始输入", "IP地址", "关联域名", "主域名及备案信息"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header).fill = HEADER_FILL
    
    row_idx = 2
    total_rows = len(results)
    processed_rows = 0
    
    print(f"\n📊 正在导出Excel，共 {total_rows} 条记录...")
    print("-" * 50)
    
    for target, data in results.items():
        # 检查是否收到中断信号
        if save_state_flag:
            print(f"\n⚠️ 导出被中断，已处理 {processed_rows}/{total_rows} 条记录")
            break
        
        if not data:
            ws.append([target, None, "解析失败", ""])
            processed_rows += 1
            continue
            
        domains = data['domains']
        original_host = urlparse(target).hostname
        
        highlight = any(
            (data['ip'] in d) or 
            (original_host and original_host in d)
            for d in domains if isinstance(d, str)
        )
        
        # 提取所有主域名并去重（单条IP内部去重），同时查询ICP备案
        main_domains_with_icp = []
        if domains:
            seen_domains = set()  # 单条IP主域名去重
            for d in domains:
                if isinstance(d, str):
                    main_domain = extract_main_domain(d)
                    if main_domain and main_domain not in seen_domains:
                        seen_domains.add(main_domain)
                        
                        # 查询ICP备案信息（使用缓存避免重复查询）
                        icp_info = query_icp_with_cache(main_domain, use_proxy, icp_cache)
                        if icp_info is None:  # 被中断
                            break
                        if icp_info:
                            main_domains_with_icp.append(f"{main_domain}-{icp_info}")
                        else:
                            main_domains_with_icp.append(main_domain)
                        # 添加延时避免请求过快
                        time.sleep(0.5)
        
        row = [
            target,
            data['ip'],
            "\n".join(domains) if domains else "无结果",
            "\n".join(main_domains_with_icp) if main_domains_with_icp else ""
        ]
        ws.append(row)
        
        if highlight:
            ws.cell(row=row_idx, column=3).fill = HIGHLIGHT_FILL
        
        row_idx += 1
        processed_rows += 1
        
        # 显示进度
        if processed_rows % 10 == 0 or processed_rows == total_rows:
            percent = (processed_rows / total_rows) * 100
            print(f"\r导出进度: {processed_rows}/{total_rows} ({percent:.1f}%)", end='', flush=True)
    
    print()  # 换行
    
    for col in ws.columns:
        max_len = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2
    
    wb.save(filename)
    print(f"✅ 结果已保存到: {filename}")
    
    return processed_rows

def export_results(results, filename, use_proxy=True):
    """导出Excel（兼容旧版本调用）"""
    export_results_with_progress(results, filename, use_proxy, None)

def main(argv):
    print_banner()
    targets = []
    output = "results.xlsx"
    proxy = None

    try:
        opts, args = getopt.getopt(argv, "hu:l:o:p:", ["help", "url=", "list=", "output=", "proxy="])
    except getopt.GetoptError:
        print("参数错误！使用 -h 查看帮助")
        sys.exit(2)

    for opt, arg in opts:
        if opt == '-h':
            print(f"Usage: {sys.argv[0]} [-u URL/IP] [-l FILE] [-o FILE] [-p PROXY]")
            print(f"  -u, --url     单个目标URL或IP")
            print(f"  -l, --list    包含目标列表的文件")
            print(f"  -o, --output  输出文件名 (默认: results.xlsx)")
            print(f"  -p, --proxy   代理地址，例如: http://127.0.0.1:7890")
            sys.exit()
        elif opt in ("-u", "--url"):
            targets.append(arg)
        elif opt in ("-l", "--list"):
            try:
                with open(arg, 'r') as f:
                    targets.extend(line.strip() for line in f if line.strip())
            except FileNotFoundError:
                print(f"文件不存在: {arg}")
                sys.exit(1)
        elif opt in ("-o", "--output"):
            output = arg
        elif opt in ("-p", "--proxy"):
            proxy = arg

    if not targets:
        print("请指定目标(-u/-l)")
        sys.exit(1)

    # 检查是否有未完成的进度
    global current_state, save_state_flag
    saved_state = load_state()
    icp_cache = {}  # ICP查询缓存
    export_completed = False  # 导出是否完成
    
    if saved_state:
        print(f"\n📂 发现未完成的进度:")
        print(f"   已处理: {saved_state.get('processed_count', 0)}/{saved_state.get('total_targets', 0)}")
        print(f"   输出文件: {saved_state.get('output', 'results.xlsx')}")
        user_input = input("是否继续上次的进度? (yes/no): ").strip().lower()
        if user_input == 'yes':
            # 恢复上次的进度
            results = {k: v for k, v in saved_state.get('results', {}).items()}
            processed_targets = set(saved_state.get('processed_targets', []))
            output = saved_state.get('output', output)
            icp_cache = saved_state.get('icp_cache', {})
            export_completed = saved_state.get('export_completed', False)
            # 过滤掉已处理的目标
            targets = [t for t in targets if t not in processed_targets]
            print(f"   将继续处理剩余 {len(targets)} 个目标")
            if icp_cache:
                print(f"   已缓存 {len(icp_cache)} 个域名的ICP信息")
            if not export_completed and len(targets) == 0:
                print(f"   IP查询已完成，将继续导出Excel")
        else:
            clear_state()
            results = {}
    else:
        results = {}

    # 配置代理
    global PROXY
    if proxy:
        # 自动检测代理协议，支持 http/https/socks5
        if proxy.startswith('socks5://'):
            PROXY = {
                'http': proxy,
                'https': proxy
            }
        elif proxy.startswith('http://') or proxy.startswith('https://'):
            PROXY = {
                'http': proxy,
                'https': proxy
            }
        else:
            # 默认添加 http:// 前缀
            proxy = f"http://{proxy}"
            PROXY = {
                'http': proxy,
                'https': proxy
            }
        print(f"\n🌐 使用代理: {proxy}")
        print(f"💡 提示: 如果使用 FlClash/V2RayN 等工具，可尝试 socks5://127.0.0.1:7890")
    else:
        PROXY = {'http': None, 'https': None}
        print(f"\n🌐 不使用代理")

    print(f"\n🔍 开始处理 {len(targets)} 个目标...")
    print("-" * 50)
    
    # 初始化查询计数器
    global query_counter
    query_counter['current'] = 0
    query_counter['total'] = len(targets)
    
    processed_count = 0
    use_proxy = proxy is not None
    total_targets = len(targets)
    processed_targets = list(results.keys())  # 已处理的目标
    
    # 更新全局状态
    current_state = {
        'output': output,
        'total_targets': total_targets + len(processed_targets),
        'processed_count': len(processed_targets),
        'results': results,
        'processed_targets': processed_targets,
        'icp_cache': icp_cache,
        'export_completed': export_completed
    }
    
    def print_progress(current, total, target, status="处理中"):
        """打印进度条"""
        percent = (current / total) * 100
        bar_length = 30
        filled = int(bar_length * current / total)
        bar = '█' * filled + '░' * (bar_length - filled)
        print(f"\r[{bar}] {current}/{total} ({percent:.1f}%) | {status} | 当前: {target[:30]}{'...' if len(target) > 30 else ''}", end='', flush=True)
    
    try:
        with ThreadPoolExecutor(max_workers=3) as executor:
            futures = {executor.submit(process_target, t, use_proxy): t for t in targets}
            for future in futures:
                # 检查是否收到中断信号
                if save_state_flag:
                    break
                    
                target = futures[future]
                result = future.result()
                # process_target 返回 (target, data) 元组
                if isinstance(result, tuple) and len(result) == 2:
                    _, data = result
                else:
                    data = result
                results[target] = data
                processed_targets.append(target)
                processed_count += 1
                
                # 更新状态
                current_state['processed_count'] = len(processed_targets)
                current_state['results'] = results
                current_state['processed_targets'] = processed_targets
                current_state['icp_cache'] = icp_cache
                current_state['export_completed'] = export_completed
                
                # 显示进度
                status = "✓" if data and isinstance(data, dict) and data.get('domains') else "✗"
                print_progress(processed_count, total_targets, target, status)
                
                # 每个目标处理后添加延迟，避免请求过快
                time.sleep(random.uniform(2, 4))

        
        # 如果是因为中断退出的，保存状态
        if save_state_flag:
            current_state['icp_cache'] = icp_cache
            current_state['export_completed'] = export_completed
            save_state()
            print("\n\n✅ 进度已保存，下次运行可继续")
            return
        else:
            # 正常完成IP查询阶段
            pass
            
    except KeyboardInterrupt:
        current_state['icp_cache'] = icp_cache
        current_state['export_completed'] = export_completed
        save_state()
        print("\n\n✅ 进度已保存，下次运行可继续")
        return
    
    # 导出Excel阶段
    if not export_completed:
        try:
            processed_rows = export_results_with_progress(results, output, use_proxy, icp_cache)
            # 如果导出完成（没有被中断）
            if processed_rows == len(results) and not save_state_flag:
                export_completed = True
                clear_state()
                print("\n🎉 所有任务已完成！")
            else:
                # 导出被中断，保存状态
                current_state['icp_cache'] = icp_cache
                current_state['export_completed'] = False
                save_state()
                print("\n\n✅ 导出进度已保存，下次运行可继续")
        except KeyboardInterrupt:
            current_state['icp_cache'] = icp_cache
            current_state['export_completed'] = False
            save_state()
            print("\n\n✅ 导出进度已保存，下次运行可继续")
    else:
        print("\n✅ Excel导出已完成，跳过导出阶段")
        clear_state()

if __name__ == "__main__":
    main(sys.argv[1:])
